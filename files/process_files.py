import os
import pandas as pd
from .models import Files
import django_eventstream
from openpyxl import load_workbook


def send_progress_update(file_id, progress, status):
    django_eventstream.send_event(
        "file-" + str(file_id),
        "message",
        {
            "file_id": str(file_id),
            "progress": progress,
            "status": status
        }
    )


def process_file(file_id):
    file_object = Files.objects.get(id=file_id)
    file_object.status = "processing"
    file_object.progress = 0
    file_object.save()

    print()
    print()
    print()
    print()
    print()
    print()
    file_path = file_object.file.path
    print(file_path)
    print()
    print()
    print()
    print()
    print()
    parsed_data = []

    try:
      
        if file_path.endswith(".csv"):
           
            total_lines = sum(1 for _ in open(file_path, "r", encoding="utf-8", errors="ignore"))
            processed_lines = 0

        
            for chunk in pd.read_csv(file_path, chunksize=5000, encoding="utf-8", on_bad_lines="skip"):
                processed_lines += len(chunk)
                progress = int((processed_lines / total_lines) * 100)


         
                parsed_data.extend(chunk.to_dict(orient="records"))

             
                file_object.progress = progress
                file_object.save()
                send_progress_update(file_id, progress, "processing")

        elif file_path.endswith(".xlsx"):
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            total_rows = ws.max_row
            processed_rows = 0

            rows_list = []
            for row in ws.iter_rows(values_only=True):
                processed_rows += 1
                progress = int((processed_rows / total_rows) * 100)

                if len(rows_list) < 100:
                    rows_list.append(row)

                file_object.progress = progress
                file_object.save()
                send_progress_update(file_id, progress, "processing")

        
            headers = rows_list[0]
            parsed_data = [dict(zip(headers, row)) for row in rows_list[1:]]

        else:
            file_object.status = "failed"
            file_object.parsed_content = {"error": "Unsupported file format"}
            file_object.save()
            send_progress_update(file_id, 0, "failed")
            return

   
        file_object.parsed_content = parsed_data
        file_object.status = "ready"
        file_object.progress = 100
        file_object.save()
        send_progress_update(file_id, 100, "completed")

    except Exception as e:
        file_object.status = "failed"
        file_object.progress = 0
        file_object.parsed_content = {"error": "file type not supprted"}
        file_object.save()
        send_progress_update(file_id, 0, "failed")
